#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Extract words from a single consolidated Excel file (all.xlsx)
Simplified version without POS tagging - all words treated equally
"""

import sys
import os
from openpyxl import load_workbook
import pickle

def extract_words_from_excel(filepath):
    """Extract all words from the consolidated Excel file"""
    print(f"\n{'='*70}")
    print(f"📊 Processing: {os.path.basename(filepath)}")
    print(f"{'='*70}\n")
    
    words = set()  # Using set for unique words
    
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active  # Get first sheet
        
        max_row = ws.max_row
        max_col = ws.max_column
        
        print(f"   Dimensions: {max_row} rows × {max_col} columns")
        
        word_count = 0
        
        # Skip header row, start from row 3
        for row_idx in range(3, max_row + 1):
            for col_idx in range(2, max_col + 1):  # Skip first column (serial number)
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                if value and isinstance(value, str):
                    value = value.strip()
                    
                    # Skip empty strings and numbers
                    if value and not value.replace('.', '').replace(',', '').isdigit():
                        # Extract the base word (before space or parenthesis)
                        word = value.split()[0].split('(')[0].strip()
                        
                        if word and len(word) > 1:
                            words.add(word)
                            word_count += 1
            
            # Progress indicator
            if row_idx % 1000 == 0:
                print(f"   Processed {row_idx}/{max_row} rows...")
        
        wb.close()
        
        unique_words = len(words)
        print(f"\n   ✅ Extracted {word_count} word occurrences")
        print(f"   ✅ Unique words: {unique_words}")
        
        return words
        
    except Exception as e:
        print(f"   ❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return set()

def main():
    check_pos_dir = r"d:\NLP-Based-Kannada-Spell-Correction-System\check_pos"
    
    # Single consolidated Excel file
    excel_file = "all.xlsx"
    filepath = os.path.join(check_pos_dir, excel_file)
    
    print("\n" + "="*70)
    print("📚 BUILDING EXTENDED DICTIONARY FROM CONSOLIDATED EXCEL FILE")
    print("="*70)
    
    if not os.path.exists(filepath):
        print(f"\n❌ ERROR: File not found: {filepath}")
        print("Please ensure 'all.xlsx' exists in the check_pos folder")
        return
    
    # Extract all words (no POS categorization)
    all_words = extract_words_from_excel(filepath)
    
    # Save to pickle file
    output_file = "extended_dictionary.pkl"
    output_path = os.path.join(os.path.dirname(__file__), output_file)
    
    print(f"\n{'='*70}")
    print("💾 SAVING EXTENDED DICTIONARY")
    print(f"{'='*70}\n")
    
    with open(output_path, 'wb') as f:
        pickle.dump(all_words, f)
    
    print(f"   ✅ Saved to: {output_file}")
    print(f"   ✅ Total unique words: {len(all_words)}")
    
    # Test if specific words are included
    print(f"\n{'='*70}")
    print("🔍 CHECKING FOR SPECIFIC WORDS")
    print(f"{'='*70}\n")
    
    test_words = ['avaralli', 'ivaralli', 'ivaru', 'kannada', 'bareyalu']
    
    for word in test_words:
        if word in all_words:
            print(f"   ✅ '{word}' found in dictionary")
        else:
            print(f"   ❌ '{word}' NOT found in dictionary")
    
    print(f"\n{'='*70}")
    print("✅ DONE - Dictionary file created!")
    print(f"{'='*70}\n")

if __name__ == "__main__":
    main()
